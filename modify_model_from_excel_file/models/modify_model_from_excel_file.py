# -*- coding: utf-8 -*-

import base64
import uuid
from ast import literal_eval
from io import BytesIO
from tempfile import TemporaryFile
from zipfile import BadZipFile

import xlrd
import xlwt
from odoo import Command, _, api, fields, models
from odoo.addons.base_import.models.base_import import ImportValidationError
from odoo.exceptions import ValidationError
from odoo.models import fix_import_export_id_paths
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

FROM_EXCEL_FILE_MIXIN_MODEL_NAME = "modify.model_fields.from.excel_file"
IMPORT_FILE_MODEL_NAME = "xlsx.import"
IMPORT_TEMPLATE_MODEL_NAME = "xlsx.template"
UNSUPPORTED_FILE_MSG = "Unsupported file format, this import option only supports excel files (*.xls, *.xlsx)."


class ModifyModelFieldsFromExcelFileMixin(models.AbstractModel):
    _name = FROM_EXCEL_FILE_MIXIN_MODEL_NAME
    _description = "Mixin to fill any field that you may have from excel files."

    import_file = fields.Binary(string="Import File (*.xls, *.xlsx)")
    import_template_id = fields.Many2one(
        comodel_name=IMPORT_TEMPLATE_MODEL_NAME,
        string="Import Template",
    )
    _default_sheet_name = "Items"
    preview_file_content = fields.Boolean(
        string="Preview file content?",
        help="Mark if you want to fill the related fields instead of waiting the saving process.",
    )

    @api.onchange("import_file")
    def onchange_import_file(self):
        self.preview_file_content = None
        if self.import_file:
            return self._check_excel_file_content(self.import_file)

    @api.constrains("import_file")
    def constrains_import_file(self):
        for mixin in self:
            if mixin.import_file:
                self._check_excel_file_content(mixin.import_file, for_onchange=False)

    @api.onchange("preview_file_content")
    def onchange_preview_file_content(self):
        if self.preview_file_content:
            validation_error_msg_header = _("Validation Error")
            bad_result_dict = {
                "warning": {"title": validation_error_msg_header},
                "value": {"preview_file_content": None},
            }
            if not self.import_file:
                bad_result_dict["warning"]["message"] = _("Verify specified file.")
                return bad_result_dict
            if not self.import_template_id:
                bad_result_dict["warning"]["message"] = _("Verify specified template file.")
                return bad_result_dict
            data_dict = literal_eval(self.import_template_id.instruction.strip())
            if not data_dict.get("__IMPORT__", None):
                bad_result_dict["warning"]["message"] = (
                    _("An import section must exists in template %s") % self.import_template_id.name
                )
                return bad_result_dict
            data_dict = data_dict["__IMPORT__"]
            for s, _sv in data_dict.copy().items():
                for f, _fv in data_dict[s].copy().items():
                    if "_NODEL_" in f:
                        new_fv = data_dict[s].pop(f)
                        data_dict[s][f.replace("_NODEL_", "")] = new_fv

            # The beginning of what we do when creating or updating.
            import_excel_file = self._check_excel_file_content(
                self.import_file, for_onchange=False, convert_to_xls=True
            )
            CustomImportModel = self.env[IMPORT_FILE_MODEL_NAME]
            import_excel_file = self._check_excel_file_content(
                import_excel_file, for_onchange=False, convert_to_xls=True
            )
            b64bytesencoded = base64.encodebytes(base64.b64decode(import_excel_file))

            # The same that is done in `_import_record_data` method from excel_import_export/models/xlsx_import.py
            # to prepare the data to be imported.
            header_fields = []
            decoded_data = base64.decodebytes(b64bytesencoded)
            wb = xlrd.open_workbook(file_contents=decoded_data)
            out_wb = xlwt.Workbook()
            out_st = out_wb.add_sheet(self._default_sheet_name)
            xml_id = "{}.{}".format("xls", uuid.uuid4())
            out_st.write(0, 0, "id")  # id and xml_id on first column
            out_st.write(1, 0, xml_id)
            header_fields.append("id")
            CustomImportModel._process_worksheet(wb, out_wb, out_st, self._name, data_dict, header_fields)

            content = BytesIO()
            out_wb.save(content)
            content.seek(0)  # Set index to 0, and start reading
            xls_file = content.read()

            BaseImportModel = self.env["base_import.import"]
            import_options = {
                "headers": True,
                "advanced": True,
                "keep_matches": False,
                "encoding": "",
                "separator": "",
                "quoting": '"',
                "date_format": "%Y-%m-%d",
                "datetime_format": "%Y-%m-%d %H:%M:%S",
                "float_thousand_separator": ",",
                "float_decimal_separator": ".",
                "fields": [],
            }
            base_import_instance = BaseImportModel.create(
                {
                    "res_model": self._name,
                    "file": xls_file,
                    "file_type": "application/vnd.ms-excel",
                    "file_name": "temp.xls",
                }
            )
            try:
                input_file_data, import_fields = base_import_instance._convert_import_data(
                    header_fields, import_options
                )
                input_file_data = base_import_instance._parse_import_data(
                    input_file_data, import_fields, import_options
                )
            except ValueError:
                bad_result_dict["warning"]["message"] = _("Conversion of values ​​could not be done.")
                return bad_result_dict
            except ImportValidationError:
                bad_result_dict["warning"]["message"] = _("Parsing of values ​​could not be done.")
                return bad_result_dict

            CurrentModel = self.env[self._name]
            try:
                import_fields, merged_data = base_import_instance._handle_multi_mapping(import_fields, input_file_data)
                import_fields = [fix_import_export_id_paths(f) for f in import_fields]
                extracted = CurrentModel._extract_records(import_fields, merged_data)
                converted = CurrentModel._convert_records(extracted)
                _record_id, _record_xid, record_data, record_location_info = list(converted)[0]
                for rkey, rvalue in record_data.copy().items():
                    if CurrentModel._fields[rkey].type in ["one2many"]:
                        record_data[rkey] = [Command.set([])] + rvalue
                return {"value": record_data}
            except ValueError:
                bad_result_dict["warning"]["message"] = _("Parsing of values ​​could not be done.")
                return bad_result_dict

    @api.model
    def _check_excel_file_content(self, base64_data, for_onchange=True, convert_to_xls=False, b64econded=True):
        decoded_data = base64.b64decode(base64_data)
        successfull_result = {} if for_onchange else True
        try:
            xlrd.open_workbook(file_contents=decoded_data)
            return successfull_result if for_onchange else base64_data
        except xlrd.XLRDError:
            unsupported_file_msg = _(UNSUPPORTED_FILE_MSG)
            bad_result_dict = {
                "warning": {"title": "Error de validación", "message": unsupported_file_msg},
                "value": {"import_file": None},
            }
            try:
                with TemporaryFile() as xlsx_file:
                    xlsx_file.write(decoded_data)
                    xlsx_file.seek(0)
                    xlsx_workbook = load_workbook(filename=xlsx_file, data_only=True)
                    if convert_to_xls:
                        return self._convert_to_xls(xlsx_workbook, b64econded=b64econded)
                    return successfull_result
            except InvalidFileException:
                if for_onchange:
                    return bad_result_dict
                else:
                    raise ValidationError(unsupported_file_msg)
            except BadZipFile:
                if for_onchange:
                    return bad_result_dict
                else:
                    raise ValidationError(unsupported_file_msg)

    @api.model
    def _convert_to_xls(self, xlsx_workbook, b64econded=True):
        xslx_active_sheet = xlsx_workbook.active
        xls_workbook = xlwt.Workbook()
        xls_active_sheet = xls_workbook.add_sheet(xslx_active_sheet.title)
        for row_index, row in enumerate(
            xslx_active_sheet.iter_rows(
                min_row=xslx_active_sheet.min_row,
                max_row=xslx_active_sheet.max_row,
                min_col=xslx_active_sheet.min_column,
                max_col=xslx_active_sheet.max_column,
            )
        ):
            for col_index, row_info in enumerate(row):
                xls_active_sheet.write(row_index, col_index, row_info.value)
        if not b64econded:
            return xls_workbook
        else:
            content = BytesIO()
            xls_workbook.save(content)
            content.seek(0)  # Set index to 0, and start reading
            return base64.b64encode(content.read())

    @api.model
    def create(self, vals_list):
        created = super().create(vals_list)
        import_excel_file = vals_list.get("import_file", None)
        preview_file_content = vals_list.get("preview_file_content", None)
        if import_excel_file and not preview_file_content:
            created.perform_import(vals_list["import_template_id"], import_excel_file)
        return created

    def write(self, vals):
        result = super().write(vals)
        import_excel_file = vals.get("import_file", None)
        preview_file_content = vals.get("preview_file_content", None)
        if result and import_excel_file and not preview_file_content:
            for mixin in self:
                mixin.perform_import(mixin.import_template_id.id, import_excel_file)
        return result

    def perform_import(self, import_template_id, import_excel_file):
        self.ensure_one()

        CustomImportModel = self.env[IMPORT_FILE_MODEL_NAME]
        import_excel_file_template = self.env[IMPORT_TEMPLATE_MODEL_NAME].browse(import_template_id)
        import_excel_file = self._check_excel_file_content(import_excel_file, for_onchange=False, convert_to_xls=True)
        CustomImportModel.import_xlsx(
            base64.encodebytes(base64.b64decode(import_excel_file)),
            import_excel_file_template,
            res_model=self._name,
            res_id=self.id,
        )
